import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

# Load the two Excel files into pandas DataFrames
file1= 'C:/Vivek/pre.xlsx' 
file2= 'C:/Vivek/post.xlsx'

df1 = pd.read_excel(file1, dtype=str)  # Read as string to preserve any non-date formatting
df2 = pd.read_excel(file2, dtype=str)  # Read as string to preserve any non-date formatting

# Ensure both DataFrames have the same columns
if not df1.columns.equals(df2.columns):
    raise ValueError("The columns of both Excel files are different!")

# Prepare a list to store the differences
differences = []

# Load the Excel files using openpyxl to check for number/date formats
wb1 = openpyxl.load_workbook(file1)
wb2 = openpyxl.load_workbook(file2)
ws1 = wb1.active
ws2 = wb2.active

# Create a merged list to store merged data
merged_data = []

# Iterate over rows, comparing each one
min_length = min(len(df1), len(df2))
max_length = max(len(df1), len(df2))

# Loop through rows, comparing each cell for differences
for idx in range(max_length):
    row1 = df1.iloc[idx] if idx < len(df1) else pd.Series([None] * len(df1.columns), index=df1.columns)
    row2 = df2.iloc[idx] if idx < len(df2) else pd.Series([None] * len(df2.columns), index=df2.columns)

    merged_row = []

    # Loop through columns, comparing values and formats
    for col_idx, col in enumerate(df1.columns):
        file1_value = row1[col]
        file2_value = row2[col]

        # Get the corresponding cells from openpyxl for format comparison
        cell1 = ws1.cell(row=idx + 1, column=col_idx + 1)
        cell2 = ws2.cell(row=idx + 1, column=col_idx + 1)

        # Extract the raw values for comparison
        file1_raw = cell1.value
        file2_raw = cell2.value

        # Get the number/date format strings
        file1_format = cell1.number_format
        file2_format = cell2.number_format

        # Check if values are numeric or dates, and compare
        try:
            # Attempt to convert to numeric values
            file1_value_num = pd.to_numeric(file1_value, errors='raise')
            file2_value_num = pd.to_numeric(file2_value, errors='raise')
        except ValueError:
            file1_value_num = file2_value_num = None

        # Check if both values are dates
        try:
            file1_value_dt = pd.to_datetime(file1_value, errors='raise') if pd.notna(file1_value) else None
            file2_value_dt = pd.to_datetime(file2_value, errors='raise') if pd.notna(file2_value) else None
        except (ValueError, TypeError):
            file1_value_dt = file2_value_dt = None

        # If both values are numeric
        if file1_value_num is not None and file2_value_num is not None:
            if file1_value_num != file2_value_num:
                differences.append({
                    'Row Index': idx + 1,
                    'Column': col,
                    'Cell': f'{get_column_letter(col_idx + 1)}{idx + 1}',
                    'File 1 Value': file1_value,
                    'File 2 Value': file2_value,
                    'Note': 'Number value mismatch'
                })
                merged_row.append(f'**{file1_value}** (changed)')
            else:
                # Compare the number formats directly
                if file1_format != file2_format:
                    differences.append({
                        'Row Index': idx + 1,
                        'Column': col,
                        'Cell': f'{get_column_letter(col_idx + 1)}{idx + 1}',
                        'File 1 Value': file1_value,
                        'File 2 Value': file2_value,
                        'Note': 'Number format mismatch'
                    })
                    merged_row.append(f'**{file1_value}** (changed)')
                else:
                    merged_row.append(file1_value)

        # If both values are dates
        elif file1_value_dt is not None and file2_value_dt is not None:
            if file1_value_dt != file2_value_dt:
                differences.append({
                    'Row Index': idx + 1,
                    'Column': col,
                    'Cell': f'{get_column_letter(col_idx + 1)}{idx + 1}',
                    'File 1 Value': file1_value,
                    'File 2 Value': file2_value,
                    'Note': 'Date value mismatch'
                })
                merged_row.append(f'**{file1_value}** (changed)')
            else:
                # Compare the date formats directly
                if file1_format != file2_format:
                    differences.append({
                        'Row Index': idx + 1,
                        'Column': col,
                        'Cell': f'{get_column_letter(col_idx + 1)}{idx + 1}',
                        'File 1 Value': file1_value,
                        'File 2 Value': file2_value,
                        'Note': 'Date format mismatch'
                    })
                    merged_row.append(f'**{file1_value}** (changed)')
                else:
                    merged_row.append(file1_value)

        # If values are not numeric or date but are different
        elif file1_value != file2_value:
            differences.append({
                'Row Index': idx + 1,
                'Column': col,
                'Cell': f'{get_column_letter(col_idx + 1)}{idx + 1}',
                'File 1 Value': file1_value if pd.notna(file1_value) else 'N/A',
                'File 2 Value': file2_value if pd.notna(file2_value) else 'N/A',
                'Note': 'Value mismatch'
            })
            merged_row.append(f'**{file1_value}** (changed)')
        else:
            merged_row.append(file1_value if pd.notna(file1_value) else file2_value)

    # Add merged row to the list
    merged_data.append(merged_row)

# Convert merged data into a DataFrame
merged_df = pd.DataFrame(merged_data, columns=df1.columns)

# Save the merged DataFrame to an Excel file
merged_report_path = 'C:/Vivek/merged_report_with_number_date_check_v5.xlsx'
merged_df.to_excel(merged_report_path, index=False)

# Open the merged report using openpyxl to highlight differences
wb = openpyxl.load_workbook(merged_report_path)
ws = wb.active

# Define a fill color for highlighting the differences (yellow)
highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# Highlight the cells with differences
for diff in differences:
    row_idx = diff['Row Index']
    col = diff['Column']
    col_idx = df1.columns.get_loc(col) + 1  # Openpyxl uses 1-based index for columns
    cell = ws.cell(row=row_idx + 1, column=col_idx)  # Row index is 1-based
    cell.fill = highlight_fill

# Save the final merged report with highlighted differences
highlighted_report_path = 'C:/Vivek/highlighted_merged_report_with_number_date_check_v5.xlsx'
wb.save(highlighted_report_path)

print(f"Highlighted merged report saved to {highlighted_report_path}")

# Generate a difference report for the mismatches with cell references
if differences:
    differences_df = pd.DataFrame(differences)
    difference_report_path = 'C:/Vivek/difference_report_with_cells_number_date_check_v5.xlsx'
    differences_df.to_excel(difference_report_path, index=False)
    print(f'Difference report saved to {difference_report_path}')
else:
    print("No differences found.")
