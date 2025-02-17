import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill

# Load the two Excel files into pandas DataFrames
file1= 'C:/Vivek/pre.xlsx' 
file2= 'C:/Vivek/post.xlsx'

df1 = pd.read_excel(file1)
df2 = pd.read_excel(file2)

# Get the maximum number of rows and columns to iterate over both DataFrames
max_rows = max(len(df1), len(df2))
max_columns = max(len(df1.columns), len(df2.columns))

# Prepare a list to store the differences
differences = []

# Create a new DataFrame for merged output
merged_df = pd.DataFrame(index=range(max_rows), columns=range(max_columns))

# Iterate over the cells of both DataFrames to merge and compare
for row_idx in range(max_rows):
    for col_idx in range(max_columns):
        # Handle missing rows and columns by assigning 'N/A' for missing cells
        file1_value = df1.iloc[row_idx, col_idx] if row_idx < len(df1) and col_idx < len(df1.columns) else 'N/A'
        file2_value = df2.iloc[row_idx, col_idx] if row_idx < len(df2) and col_idx < len(df2.columns) else 'N/A'
        
        # Add the value to the merged DataFrame
        merged_df.iloc[row_idx, col_idx] = file1_value
        
        # If the values in the two files are different
        if file1_value != file2_value:
            differences.append({
                'Row Index': row_idx + 1,  # 1-based index for human readability
                'Column Index': col_idx + 1,  # 1-based index for human readability
                'File 1 Value': file1_value,
                'File 2 Value': file2_value
            })

# Save the merged DataFrame to a new Excel file
merged_report_path = 'C:/Vivek/merged_report.xlsx'
merged_df.to_excel(merged_report_path, index=False)

# Open the merged report using openpyxl to highlight differences
wb = openpyxl.load_workbook(merged_report_path)
ws = wb.active

# Define a fill color for highlighting the differences (yellow)
highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# Highlight the cells with differences
for diff in differences:
    row_idx = diff['Row Index']
    col_idx = diff['Column Index']
    # Highlight the corresponding cell in the merged file
    cell = ws.cell(row=row_idx, column=col_idx)
    cell.fill = highlight_fill

# Save the final merged report with highlighted differences
highlighted_report_path = 'C:/Vivek/highlighted_merged_report.xlsx'
wb.save(highlighted_report_path)

print(f"Highlighted merged report saved to {highlighted_report_path}")

# Generate a difference report for the mismatches
if differences:
    differences_df = pd.DataFrame(differences)
    difference_report_path = 'C:/Vivek/difference_report.xlsx'
    differences_df.to_excel(difference_report_path, index=False)
    print(f'Difference report saved to {difference_report_path}')
else:
    print("No differences found.")
