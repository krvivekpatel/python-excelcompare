import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

# Load the two Excel files into pandas DataFrames
file1= 'C:/Vivek/pre.xlsx' 
file2= 'C:/Vivek/post.xlsx'

df1 = pd.read_excel(file1)
df2 = pd.read_excel(file2)

# Ensure both DataFrames have the same columns
if not df1.columns.equals(df2.columns):
    raise ValueError("The columns of both Excel files are different!")

# Prepare a list to store the differences
differences = []

# Prepare a merged DataFrame to store the merged rows
merged_data = []

# Compare rows in both DataFrames
min_length = min(len(df1), len(df2))
max_length = max(len(df1), len(df2))

# Create a merged DataFrame row by row
for idx in range(max_length):
    row1 = df1.iloc[idx] if idx < len(df1) else pd.Series([None] * len(df1.columns), index=df1.columns)
    row2 = df2.iloc[idx] if idx < len(df2) else pd.Series([None] * len(df2.columns), index=df2.columns)

    merged_row = []

    # Compare each column in the row
    for col_idx, col in enumerate(df1.columns):
        file1_value = row1[col]
        file2_value = row2[col]

        # If the values are different or missing
        if pd.isna(file1_value) or pd.isna(file2_value) or file1_value != file2_value:
            differences.append({
                'Row Index': idx + 1,  # 1-based index
                'Column': col,
                'File 1 Value': file1_value if not pd.isna(file1_value) else 'N/A',
                'File 2 Value': file2_value if not pd.isna(file2_value) else 'N/A'
            })
            merged_row.append(f'**{file1_value}**{file2_value}** (changed)' if file1_value != file2_value else f'**{file1_value}**{file2_value}** (changed)')
        else:
            merged_row.append(file1_value if not pd.isna(file1_value) else file2_value)

    # Add merged row to the list
    merged_data.append(merged_row)

# Convert merged data into a DataFrame
merged_df = pd.DataFrame(merged_data, columns=df1.columns)

# Save the merged DataFrame to an Excel file
merged_report_path = 'C:/Vivek/merged_report.xlsx'
merged_df.to_excel(merged_report_path, index=False)

# Open the merged report using openpyxl to highlight differences
wb = openpyxl.load_workbook(merged_report_path)
ws = wb.active

# Define a fill color for highlighting the differences (yellow)
highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# Highlight the cells with differences
for diff in differences:
    row_idx = diff['Row Index']
    col = diff['Column']
    col_idx = df1.columns.get_loc(col) + 1  # Openpyxl uses 1-based index for columns
    cell = ws.cell(row=row_idx + 1, column=col_idx)  # Row index is 1-based
    cell.fill = highlight_fill

# Save the final merged report with highlighted differences
highlighted_report_path = 'C:/Vivek/highlighted_merged_report.xlsx'
wb.save(highlighted_report_path)

print(f"Highlighted merged report saved to {highlighted_report_path}")

# Generate a difference report for the mismatches
if differences:
    differences_df = pd.DataFrame(differences)
    difference_report_path = 'C:/Vivek/difference_report.xlsx'
    differences_df.to_excel(difference_report_path, index=False)
    print(f'Difference report saved to {difference_report_path}')
else:
    print("No differences found.")
